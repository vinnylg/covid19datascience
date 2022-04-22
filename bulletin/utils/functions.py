import time
# from sympy import root
import xlrd
from tqdm import tqdm
from io import RawIOBase
from contextlib import contextmanager
import openpyxl as op
import numpy as np
import pandas as pd
from unidecode import unidecode as unidecode
from datetime import date
from os.path import join

chuncker = lambda df,limit=100000,offset=0: [ (part,df.iloc[offset:offset+limit]) for part, offset in enumerate(range(0,len(df),limit)) ] 

from bulletin import root

# ------------------------------------------------------- UTILITÁRIOS
class progress_reader(RawIOBase):
    def __init__(self, zf, bar):
        self.bar = bar
        self.zf = zf

    def readinto(self, b):
        n = self.zf.readinto(b)
        self.bar.update(n=n)
        return n

@contextmanager
def load_with_progressbar():

    def my_get_sheet(self, zf, *other, **kwargs):
        with tqdm(total=zf._orig_file_size) as bar:
            sheet = _tmp(self, progress_reader(zf, bar), **kwargs)
        return sheet

    _tmp = xlrd.xlsx.X12Sheet.own_process_stream

    try:
        xlrd.xlsx.X12Sheet.own_process_stream = my_get_sheet
        yield
    finally:
        xlrd.xlsx.X12Sheet.own_process_stream = _tmp

class Timer:
    def __init__(self):
        self.__running = False

    def start(self):
        self.__start = time()
        self.__running = True

    def stop(self):
        if self.__running:
            self.__running = False
            self.__end = time()
            self.__time_elapsed = self.__end - self.__start
            return self.__time_elapsed
        else:
            print('Timer already stoped')

    def time(self):
        if not self.__running:
            time_elapsed = self.__time_elapsed
        else:
            time_elapsed = time() - self.__start

        return time_elapsed

    def ftime(self):
        if not self.__running:
            time_elapsed = self.__time_elapsed
        else:
            time_elapsed = time() - self.__start

        days, seconds = divmod(time_elapsed,60*60*24)
        hours, seconds = divmod(seconds, 60*60)
        minutes, seconds = divmod(seconds, 60)
        miliseconds = (seconds - int(seconds)) * 1000
        print(f"{int(hours):02}:{int(minutes):02}:{int(seconds):02}.{int(miliseconds):02}")

    def restart(self):
        if self.__running:
            self.stop()
        self.start()


# ------------------------------------------------------- OCUPACAO DE LEITOS
def int_float(number):
    try:
        if number % 1 == 0.0:
            return int(number)
        else:
            return number
    except:
        return number

def trim_overspace(text):
    parts = filter(lambda x: len(x) > 0,text.split(" "))
    return " ".join(parts)

def normalize_text(text):
    x = str(text).replace(".","").replace("*","").replace("\n","").replace(",","").replace("\t","").replace("''","'").replace("\"","'").upper()
    x = trim_overspace(x)
    x = unidecode(x)    
    if x == 'nan' or len(x) == 0 or x == '0' or x == 'nao_informado':
        return ''
    else:
        return x

def read_excel(pathfile):
    if pathfile.split('.')[-1] in ['xlsx', 'xlsm', 'xltx', 'xltm']:
        print(f"open {pathfile.split('.')[-1]} with openpyxl")
        book = None
    else:
        print(f"open {pathfile.split('.')[-1]} with xlrd")
        book = xlrd.open_workbook(pathfile)
                
    return book

def print_matrix(matrix,m,n):
    print("   ",end=" ")
    for j in range(n):
        print(f"{j}",end=" ")
    print("")

    print(f"   ",end=" ")
    for _ in range(n):
        print(f"-",end=" ")
    print("")

    for i in range(m):
        print(f"{i} [",end=" ")
        for j in range(n):
            print(f"{matrix[i][j]}", end=" ")
        print(f"] ({sum(matrix[i])})")
    
    print("   ",end=" ")
    for _ in range(n):
        print(f"-",end=" ")
    print("")

def find_matrixs(matrix, m, n, min_table_rows:int=2):
    matrixs = []
    init_point = [ np.inf, np.inf ]
    end_point = [ -np.inf, -np.inf ]

    for  i in range(m):
        for j in range(n):
            if matrix[i][j] == 1:
                if i < init_point[0]:
                    init_point[0] = i
                if j < init_point[1]:
                    init_point[1] = j
                if i > end_point[0]:
                    end_point[0] = i
                if j > end_point[1]:
                    end_point[1] = j

        if (sum(matrix[i]) == 0) and (init_point[0] != np.inf) and (end_point[0] - init_point[0] > min_table_rows):
            matrixs.append((tuple(init_point), tuple(end_point)))
            init_point = [ np.inf, np.inf ]
            end_point = [ -np.inf, -np.inf ]

    if init_point[0] != np.inf:
        matrixs.append((tuple(init_point), tuple(end_point)))
        init_point = [ np.inf, np.inf ]
        end_point = [ -np.inf, -np.inf ]

    return matrixs

def create_occupation_table(workbook, sheet_index):
    path = join(root, 'base')
    ibge = pd.read_csv(join(path,'ibge_mario.csv'), sep = ';', encoding='utf-8-sig')
    ibge = ibge.rename(columns = {'MUNICIPIO com e sem ACENTUAÇÃO':'municipio_normalize', 'IBGE 6': 'IBGE'})
    
    sh = workbook.sheet_by_index(sheet_index)
    matrix = [[ 0 for i in range(sh.ncols) ] for j in range(sh.nrows) ]
    for i in range(sh.nrows):
        for j in range(sh.ncols):
            cell = sh.cell(i,j)
            matrix[i][j] = 1 if cell.value != '' else 0
    matrixs_coord = find_matrixs(matrix,sh.nrows,sh.ncols,1)

    wb = op.Workbook()
    del wb['Sheet']
    for matrix in matrixs_coord:
        sheet = wb.create_sheet(title=f"table_{'_'.join(([f'{t[0]}x{t[1]}' for t in matrix]))}")
        (x_init,y_init), (x_end,y_end) = matrix 
        for i in range(x_init,x_end):
            for j in range(y_init,y_end+1):
                sheet.cell(i-x_init+1,j-y_init+1).value = sh.cell(i,j).value

    wb.save(join(path,'destination.xlsx'))

    df = pd.read_excel(join(path,'destination.xlsx'),skiprows=6,index_col=[0,1,2],header=[0,1,2,3]).fillna(0)
    for column in df.columns:
        df[column] = df[column].apply(int_float)

    df.columns = df.columns.droplevel(0).droplevel(0).droplevel(0)
    df.index = df.index.set_names(['MacroRegião', 'RS', 'Município'])
    index = pd.Index(['Hospital', 'UTI ADULTO Exist.', 'UTI ADULTO Ocup.', 'UTI ADULTO Livres', 'UTI ADULTO Tx de ocup.', 'ENFERMARIA ADULTO Exist.', 'ENFERMARIA ADULTO Ocup.', 'ENFERMARIA ADULTO Livres', 'ENFERMARIA ADULTO Tx de ocup.', 'UTI PEDIÁTRICO Exist.', 'UTI PEDIÁTRICO Ocup.', 'UTI PEDIÁTRICO Livres', 'UTI PEDIÁTRICO Tx de ocup.', 'ENFERMARIA PEDIÁTRICO Exist.', 'ENFERMARIA PEDIÁTRICO Ocup.', 'ENFERMARIA PEDIÁTRICO Livres', 'ENFERMARIA PEDIÁTRICO Tx de ocup.', 'UTI ADULTO SUS', 'enf ADULTO SUS', 'UTI PEDRIÁTRICO SUS', 'enf PEDIÁTRICO SUS'])
    # print(df.columns)
    # print(index)
    df.columns = index
    df = df.reset_index()
    df = df.drop(index = df.loc[df['MacroRegião'].str.contains('TOTAL')].index)
    df['municipio_normalize'] = df['Município'].apply(lambda row: normalize_text(row))
    ibge['municipio_normalize'] = ibge['municipio_normalize'].apply(lambda row: normalize_text(row))
    ibge['IBGE'] = ibge['IBGE'].astype('str')
    df = df.merge(how = 'left', on = 'municipio_normalize', right = ibge)
    df = df[['IBGE', 'MacroRegião', 'RS', 'Município', 'Hospital', 'UTI ADULTO Exist.', 'UTI ADULTO Ocup.', 'UTI ADULTO Livres', 'UTI ADULTO Tx de ocup.', 'ENFERMARIA ADULTO Exist.', 'ENFERMARIA ADULTO Ocup.', 'ENFERMARIA ADULTO Livres', 'ENFERMARIA ADULTO Tx de ocup.', 'UTI PEDIÁTRICO Exist.', 'UTI PEDIÁTRICO Ocup.', 'UTI PEDIÁTRICO Livres', 'UTI PEDIÁTRICO Tx de ocup.', 'ENFERMARIA PEDIÁTRICO Exist.', 'ENFERMARIA PEDIÁTRICO Ocup.', 'ENFERMARIA PEDIÁTRICO Livres', 'ENFERMARIA PEDIÁTRICO Tx de ocup.', 'UTI ADULTO SUS', 'enf ADULTO SUS', 'UTI PEDRIÁTRICO SUS', 'enf PEDIÁTRICO SUS']]

    df = df.drop_duplicates(['Hospital', 'Município'])
    df.loc[df['UTI PEDIÁTRICO Tx de ocup.'] == ' ', 'UTI PEDIÁTRICO Tx de ocup.'] = 0
    df['UTI ADULTO Tx de ocup.'] = df['UTI ADULTO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['UTI ADULTO Tx de ocup.']), sep = '')
    df['ENFERMARIA ADULTO Tx de ocup.'] =  df['ENFERMARIA ADULTO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['ENFERMARIA ADULTO Tx de ocup.']), sep = '')
    df['UTI PEDIÁTRICO Tx de ocup.'] = df['UTI PEDIÁTRICO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['UTI PEDIÁTRICO Tx de ocup.']), sep = '')
    df['ENFERMARIA PEDIÁTRICO Tx de ocup.'] = df['ENFERMARIA PEDIÁTRICO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['ENFERMARIA PEDIÁTRICO Tx de ocup.']), sep = '')
    
    return df


# ------------------------------------------------------- FAIXAS ETÁRIAS
def faixa_etaria_linhas(casos, file_name, faixa_etaria, faixa_etaria_labels):
    faixa_etaria_index = dict(enumerate(faixa_etaria_labels))
    index_casos = np.digitize(casos['idade_original'], faixa_etaria, right=False)

    casos['faixa_etaria'] = [ index for index in index_casos ]

    casos['mes_caso'] = casos.apply(lambda row: row['data_diagnostico'].strftime('%m'), axis=1)
    casos['ano_caso'] = casos.apply(lambda row: row['data_diagnostico'].strftime('%Y'), axis=1)


    obitos = casos.loc[ casos['data_obito'].notnull() ].copy()
    obitos['mes_obito'] = obitos.apply(lambda row: row['data_obito'].strftime('%m'), axis=1)
    obitos['ano_obito'] = obitos.apply(lambda row: row['data_obito'].strftime('%Y'), axis=1)



    faixa_etaria_casos = casos.groupby(by=['ano_caso','mes_caso','faixa_etaria'])[['sexo','ibge_res_pr']].count().rename(columns={'sexo':'0','ibge_res_pr':'1'})
    faixa_etaria_casos = faixa_etaria_casos.unstack()
    faixa_etaria_casos = faixa_etaria_casos.swaplevel(1,0,1).sort_index(1).rename(columns={'0':'qtde','1':'%'})

    faixa_etaria_casos.columns = pd.MultiIndex.from_tuples(faixa_etaria_casos.columns)
    faixa_etaria_casos['total'] = faixa_etaria_casos.xs('qtde',level=1, axis=1).sum(axis=1)


    faixa_etaria_casos = faixa_etaria_casos.sort_index()

    for ano_mes in faixa_etaria_casos.index:
        for faixa_etaria, tipo in faixa_etaria_casos.columns:
            if tipo == '%':
                faixa_etaria_casos.loc[ano_mes,(faixa_etaria,'%')] = faixa_etaria_casos.loc[ano_mes,(faixa_etaria,'%')] / faixa_etaria_casos.loc[ano_mes,'total'].values

    faixa_etaria_casos.columns = faixa_etaria_casos.columns.set_levels([ faixa_etaria_index[index] if isinstance(index,int) else index for index in faixa_etaria_casos.columns.levels[0]],0,False)





    faixa_etaria_obitos = obitos.groupby(by=['ano_obito','mes_obito','faixa_etaria'])[['sexo','ibge_res_pr']].count().rename(columns={'sexo':'0','ibge_res_pr':'1'})
    faixa_etaria_obitos = faixa_etaria_obitos.unstack()
    faixa_etaria_obitos = faixa_etaria_obitos.swaplevel(1,0,1).sort_index(1).rename(columns={'0':'qtde','1':'%'})

    faixa_etaria_obitos.columns = pd.MultiIndex.from_tuples(faixa_etaria_obitos.columns)
    faixa_etaria_obitos['total'] = faixa_etaria_obitos.xs('qtde',level=1, axis=1).sum(axis=1)

    faixa_etaria_obitos = faixa_etaria_obitos.sort_index()

    for ano_mes in faixa_etaria_obitos.index:
        for faixa_etaria, tipo in faixa_etaria_obitos.columns:
            if tipo == '%':
                faixa_etaria_obitos.loc[ano_mes,(faixa_etaria,'%')] = faixa_etaria_obitos.loc[ano_mes,(faixa_etaria,'%')] / faixa_etaria_obitos.loc[ano_mes,'total'].values
                
    faixa_etaria_obitos.columns = faixa_etaria_obitos.columns.set_levels([ faixa_etaria_index[index] if isinstance(index,int) else index for index in faixa_etaria_obitos.columns.levels[0]],0,False)


    writer = pd.ExcelWriter(file_name,
                            engine='xlsxwriter',
                            datetime_format='dd/mm/yyyy',
                            date_format='dd/mm/yyyy')

    workbook = writer.book

    faixa_etaria_casos.to_excel(writer,sheet_name=f"faixa_etaria_casos")
    worksheet = writer.sheets[f"faixa_etaria_casos"]


    faixa_etaria_obitos.to_excel(writer,sheet_name=f"faixa_etaria_obitos")
    worksheet = writer.sheets[f"faixa_etaria_obitos"]

    writer.save()

def faixa_etaria_colunas(casos, file_name, faixa_etaria, faixa_etaria_labels):
    faixa_etaria_index = dict(enumerate(faixa_etaria_labels))
    index_casos = np.digitize(casos['idade_original'], faixa_etaria, right=False)

    casos['faixa_etaria'] = [ index for index in index_casos ]

    casos['mes_caso'] = casos.apply(lambda row: row['data_diagnostico'].strftime('%m'), axis=1)
    casos['ano_caso'] = casos.apply(lambda row: row['data_diagnostico'].strftime('%Y'), axis=1)

    obitos = casos.loc[ casos['data_obito'].notnull() ].copy()
    obitos['mes_obito'] = obitos.apply(lambda row: row['data_obito'].strftime('%m'), axis=1)
    obitos['ano_obito'] = obitos.apply(lambda row: row['data_obito'].strftime('%Y'), axis=1)

    faixa_etaria_casos = casos.groupby(by=['ano_caso','mes_caso','faixa_etaria'])[['sexo']].count()
    faixa_etaria_casos = faixa_etaria_casos.unstack(0).unstack(0)
    faixa_etaria_casos = faixa_etaria_casos.swaplevel(1,0,1).sort_index(1)

    faixa_etaria_casos = faixa_etaria_casos.sort_index()

    faixa_etaria_casos.index = [ faixa_etaria_index[index] if isinstance(index,int) else index for index in faixa_etaria_casos.index]
    faixa_etaria_casos = faixa_etaria_casos.fillna(0).astype(int)




    faixa_etaria_obitos = obitos.groupby(by=['ano_obito','mes_obito','faixa_etaria'])[['sexo']].count()
    faixa_etaria_obitos = faixa_etaria_obitos.unstack(0).unstack(0)
    faixa_etaria_obitos = faixa_etaria_obitos.swaplevel(1,0,1).sort_index(1)

    faixa_etaria_obitos = faixa_etaria_obitos.sort_index()
                
    faixa_etaria_obitos.index = [ faixa_etaria_index[index] if isinstance(index,int) else index for index in faixa_etaria_obitos.index]

    faixa_etaria_obitos.columns = faixa_etaria_obitos.columns.droplevel(level = 1)
    faixa_etaria_obitos = faixa_etaria_obitos.fillna(0).astype(int)




    writer = pd.ExcelWriter(file_name,
                            engine='xlsxwriter',
                            datetime_format='dd/mm/yyyy',
                            date_format='dd/mm/yyyy')

    workbook = writer.book

    faixa_etaria_casos.to_excel(writer,sheet_name=f"faixa_etaria_casos")
    worksheet = writer.sheets[f"faixa_etaria_casos"]


    faixa_etaria_obitos.to_excel(writer,sheet_name=f"faixa_etaria_obitos")
    worksheet = writer.sheets[f"faixa_etaria_obitos"]

    writer.save()


###
# Gera um dataframe n X n sendo n as datas da ficha x cujo n não é nulo
# Calcula a diferença em dias entre cada uma das datas e salva nesse dataframe

def mean_distance_of_dates(row):
    if row.loc[row.notna()].shape != 0:
        # name = row.name
        row = row.append(pd.Series([pd.to_datetime(date.today())], index=['today']))
        # display(row)
        # row.name = name
        # display(row)
        df_distance_dates = pd.DataFrame(index=row.loc[row.notna()].index,columns=row.loc[row.notna()].index)
        
        for i in df_distance_dates.index:
            for j in df_distance_dates.columns:
                df_distance_dates.loc[i,j] = row.loc[i] - row.loc[j]

        df_mean_distance_dates = (df_distance_dates / np.timedelta64(1, 'D')).astype(int).abs().mean()
        
        idx_min_distance = df_mean_distance_dates.idxmin()
        idx_max_distance = df_mean_distance_dates.idxmax()
        
        if len(df_mean_distance_dates) > 1 and idx_min_distance == idx_max_distance:
            idx_max_distance = df_mean_distance_dates.loc[~df_mean_distance_dates.index.isin([idx_min_distance])].idxmax()
        
        data_min_distance = row[idx_min_distance]
        data_max_distance = row[idx_max_distance]
        
        serie_distance_all_dates = pd.Series(index=row.index, dtype=float)
        serie_distance_all_dates.update(df_mean_distance_dates)

        return pd.Series(data=[idx_min_distance, 
                          data_min_distance, 
                          idx_max_distance, 
                          data_max_distance, 
                          (data_min_distance-data_max_distance).days,
                          df_mean_distance_dates.count(), 
                          df_mean_distance_dates.sum(), 
                          df_mean_distance_dates.mean(), 
                          df_mean_distance_dates.std()
                         ] + serie_distance_all_dates.tolist(),
                        index=['idx_min_distance',
                               'data_min_distance', 
                               'idx_max_distance',
                               'data_max_distance',
                               'diff_min_max',
                               'n_datas_diff_mean',
                               'soma_diff_mean',
                               'media_diff_mean',
                               'desvio_padrao_diff_mean'
                              ] + [ f"{x}_diff_mean" for x in row.index ] 
                        )
    else:
        return pd.Series(data = [ None, pd.NaT, None, pd.NaT, np.NaN, np.NaN, np.NaN, np.NaN, np.NaN ] + [ np.NaN for x in row.index ] ,
                        index=['idx_min_distance',
                               'data_min_distance', 
                               'idx_max_distance',
                               'data_max_distance',
                               'diff_min_max',
                               'n_datas_diff_mean',
                               'soma_diff_mean',
                               'media_diff_mean',
                               'desvio_padrao_diff_mean'
                              ] + [ f"{x}_diff_mean" for x in row.index ] 
                        )
