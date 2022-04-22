import xlrd
import openpyxl as op
import numpy as np
import pandas as pd
from os.path import join
from bulletin import default_input, default_output

def int_float(number):
    try:
        if number % 1 == 0.0:
            return int(number)
        else:
            return number
    except:
        return number

def trim_overspace(text):
    parts = filter(lambda x: len(x) > 0,text.split(" "))
    return " ".join(parts)

def normalize_text(text):
    x = str(text).replace(".","").replace("*","").replace("\n","").replace(",","").replace("\t","").replace("''","'").replace("\"","'").upper()
    x = trim_overspace(x)
    x = unidecode(x)    
    if x == 'nan' or len(x) == 0 or x == '0' or x == 'nao_informado':
        return ''
    else:
        return x

def read_excel(pathfile):
    if pathfile.split('.')[-1] in ['xlsx', 'xlsm', 'xltx', 'xltm']:
        print(f"open {pathfile.split('.')[-1]} with openpyxl")
        book = None
    else:
        print(f"open {pathfile.split('.')[-1]} with xlrd")
        book = xlrd.open_workbook(pathfile)
                
    return book

def print_matrix(matrix,m,n):
    print("   ",end=" ")
    for j in range(n):
        print(f"{j}",end=" ")
    print("")

    print(f"   ",end=" ")
    for _ in range(n):
        print(f"-",end=" ")
    print("")

    for i in range(m):
        print(f"{i} [",end=" ")
        for j in range(n):
            print(f"{matrix[i][j]}", end=" ")
        print(f"] ({sum(matrix[i])})")
    
    print("   ",end=" ")
    for _ in range(n):
        print(f"-",end=" ")
    print("")


def find_matrixs(matrix, m, n, min_table_rows:int=2):
    matrixs = []
    init_point = [ np.inf, np.inf ]
    end_point = [ -np.inf, -np.inf ]

    for  i in range(m):
        for j in range(n):
            if matrix[i][j] == 1:
                if i < init_point[0]:
                    init_point[0] = i
                if j < init_point[1]:
                    init_point[1] = j
                if i > end_point[0]:
                    end_point[0] = i
                if j > end_point[1]:
                    end_point[1] = j

        if (sum(matrix[i]) == 0) and (init_point[0] != np.inf) and (end_point[0] - init_point[0] > min_table_rows):
            matrixs.append((tuple(init_point), tuple(end_point)))
            init_point = [ np.inf, np.inf ]
            end_point = [ -np.inf, -np.inf ]

    if init_point[0] != np.inf:
        matrixs.append((tuple(init_point), tuple(end_point)))
        init_point = [ np.inf, np.inf ]
        end_point = [ -np.inf, -np.inf ]

#     for i,m in enumerate(matrixs):
#         print(f"{i}: {m}")

    return matrixs

def normalize_workbook(workbook, sheet_index):
    sh = workbook.sheet_by_index(sheet_index)
    matrix = [[ 0 for i in range(sh.ncols) ] for j in range(sh.nrows) ]
    for i in range(sh.nrows):
        for j in range(sh.ncols):
            cell = sh.cell(i,j)
            matrix[i][j] = 1 if cell.value != '' else 0
    matrixs_coord = find_matrixs(matrix,sh.nrows,sh.ncols,1)

    wb = op.Workbook()
    del wb['Sheet']
    for matrix in matrixs_coord:
        sheet = wb.create_sheet(title=f"table_{'_'.join(([f'{t[0]}x{t[1]}' for t in matrix]))}")
        (x_init,y_init), (x_end,y_end) = matrix 
        for i in range(x_init,x_end):
            for j in range(y_init,y_end+1):
                sheet.cell(i-x_init+1,j-y_init+1).value = sh.cell(i,j).value

    wb.save(join(default_input, 'ocupacao_leitos','destination.xlsx'))

def create_occupation_table(workbook, sheet_index, sheet_name = None):
    normalize_workbook(workbook, sheet_index)
    df = pd.read_excel(join(default_input, 'ocupacao_leitos','destination.xlsx'),skiprows=6,index_col=[0,1,2],header=[0,1,2,3]).fillna(0)
    for column in df.columns:
        df[column] = df[column].apply(int_float)

    df.columns = df.columns.droplevel(0).droplevel(0).droplevel(0)
    df.index = df.index.set_names(['MacroRegião', 'RS', 'Município'])
    index = pd.Index(['Hospital', 'UTI ADULTO Exist.', 'UTI ADULTO Ocup.', 'UTI ADULTO Livres', 'UTI ADULTO Tx de ocup.', 'ENFERMARIA ADULTO Exist.', 'ENFERMARIA ADULTO Ocup.', 'ENFERMARIA ADULTO Livres', 'ENFERMARIA ADULTO Tx de ocup.', 'UTI PEDIÁTRICO Exist.', 'UTI PEDIÁTRICO Ocup.', 'UTI PEDIÁTRICO Livres', 'UTI PEDIÁTRICO Tx de ocup.', 'ENFERMARIA PEDIÁTRICO Exist.', 'ENFERMARIA PEDIÁTRICO Ocup.', 'ENFERMARIA PEDIÁTRICO Livres', 'ENFERMARIA PEDIÁTRICO Tx de ocup.', 'UTI ADULTO SUS', 'enf ADULTO SUS', 'UTI PEDRIÁTRICO SUS', 'enf PEDIÁTRICO SUS'])
    # print(df.columns,index)
    df.columns = index
    df = df.reset_index()
    df = df.drop(index = df.loc[df['MacroRegião'].str.contains('TOTAL')].index)
#     df['municipio_normalize'] = df['Município'].apply(lambda row: normalize_text(row))
#     ibge['municipio_normalize'] = ibge['municipio_normalize'].apply(lambda row: normalize_text(row))
#     ibge['IBGE'] = ibge['IBGE'].astype('str')
#     df = df.merge(how = 'left', on = 'municipio_normalize', right = ibge)
#    df = df[['IBGE', 'MacroRegião', 'RS', 'Município', 'Hospital', 'UTI ADULTO Exist.', 'UTI ADULTO Ocup.', 'UTI ADULTO Livres', 'UTI ADULTO Tx de ocup.', 'ENFERMARIA ADULTO Exist.', 'ENFERMARIA ADULTO Ocup.', 'ENFERMARIA ADULTO Livres', 'ENFERMARIA ADULTO Tx de ocup.', 'UTI PEDIÁTRICO Exist.', 'UTI PEDIÁTRICO Ocup.', 'UTI PEDIÁTRICO Livres', 'UTI PEDIÁTRICO Tx de ocup.', 'ENFERMARIA PEDIÁTRICO Exist.', 'ENFERMARIA PEDIÁTRICO Ocup.', 'ENFERMARIA PEDIÁTRICO Livres', 'ENFERMARIA PEDIÁTRICO Tx de ocup.', 'UTI ADULTO SUS', 'enf ADULTO SUS', 'UTI PEDRIÁTRICO SUS', 'enf PEDIÁTRICO SUS']]

#     df = df.drop_duplicates(['Hospital', 'Município'])
    df.loc[df['UTI PEDIÁTRICO Tx de ocup.'] == ' ', 'UTI PEDIÁTRICO Tx de ocup.'] = 0
    df['UTI ADULTO Tx de ocup.'] = df['UTI ADULTO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['UTI ADULTO Tx de ocup.']), sep = '')
    df['ENFERMARIA ADULTO Tx de ocup.'] =  df['ENFERMARIA ADULTO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['ENFERMARIA ADULTO Tx de ocup.']), sep = '')
    df['UTI PEDIÁTRICO Tx de ocup.'] = df['UTI PEDIÁTRICO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['UTI PEDIÁTRICO Tx de ocup.']), sep = '')
    df['ENFERMARIA PEDIÁTRICO Tx de ocup.'] = df['ENFERMARIA PEDIÁTRICO Tx de ocup.'].mul(100).astype('int').astype('str').str.cat(list('%' for row in df['ENFERMARIA PEDIÁTRICO Tx de ocup.']), sep = '')

    if (sheet_name):
        df.to_excel(join(default_output, 'ocupacao_leitos', 'tabelas_ocupacao_leitos',f"{sheet_name}.xlsx"), index = False)
    else:
        df.to_excel(join(default_output, 'ocupacao_leitos', 'tabelas_ocupacao_leitos',f"ocupacao_leitos_{sheet_index}.xlsx"), index = False)
    
#     tabela_macro_ENFERMARIA = df.groupby([('', region)])[[('ENFERMARIA', 'Ocup.'), ('ENFERMARIA', 'Exist.')]].sum(0).reset_index()
#     tabela_macro_ENFERMARIA[('ENFERMARIA', 'Tx. ocupação')] = tabela_macro_ENFERMARIA[('ENFERMARIA', 'Ocup.')].div(tabela_macro_ENFERMARIA[('ENFERMARIA', 'Exist.')])
#     tabela_macro_ENFERMARIA.columns = tabela_macro_ENFERMARIA.columns.droplevel(0)

#     tabela_macro_UTI = df.groupby([('', region)])[[('UTI', 'Ocup.'), ('UTI', 'Exist.')]].sum(0).reset_index()
#     tabela_macro_UTI[('UTI', 'Tx. ocupação')] = tabela_macro_UTI[('UTI', 'Ocup.')].div(tabela_macro_UTI[('UTI', 'Exist.')])
#     tabela_macro_UTI.columns = tabela_macro_UTI.columns.droplevel(0)

#     tabela_macro_UTI = tabela_macro_UTI.rename(columns={'Ocup.':'Ocupados', 'Exist.':'Existentes'})
#     tabela_macro_ENFERMARIA = tabela_macro_ENFERMARIA.rename(columns={'Ocup.':'Ocupados', 'Exist.':'Existentes'})

#     tabela_macro_ENFERMARIA.to_excel(join('C:/Users/est.joaosilva/covid19datascience', 'output', 'tabelas_ocupacao_leitos',f"tabela_{region}_ENFERMARIA_{sh.name.replace(' ','_')}.xlsx"), index = False)
#     tabela_macro_UTI.to_excel(join('C:/Users/est.joaosilva/covid19datascience', 'output', 'tabelas_ocupacao_leitos',f"tabela_{region}_UTI_{sh.name.replace(' ','_')}.xlsx"), index = False)
